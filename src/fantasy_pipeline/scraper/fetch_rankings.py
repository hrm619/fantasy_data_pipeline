"""Automated fetchers for ranking sources that support public access.

Each fetcher downloads data and saves it as a CSV in the pipeline's update/ directory
with the filename pattern expected by FILE_MAPPINGS in config.py.
"""

import csv
import io
import json
import os
import re
from typing import Optional

import requests

from fantasy_pipeline.config import CURRENT_SEASON, LAST_COMPLETED_SEASON


USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"


def _extract_fp_report_config(html: str) -> dict:
    """Return the `window.FP.reportConfig` JSON object embedded in a FantasyPros report page.

    Uses ``raw_decode`` (brace matching) rather than a non-greedy regex: the blob contains
    nested braces and JSON strings, so a `(\\{.*?\\});` match can terminate early.
    """
    match = re.search(r"window\.FP\.reportConfig\s*=\s*", html)
    if not match:
        raise RuntimeError("Could not find the reportConfig JSON on the FantasyPros report page — layout changed")
    try:
        config, _ = json.JSONDecoder().raw_decode(html, match.end())
    except ValueError as exc:
        raise RuntimeError(f"Could not decode the FantasyPros reportConfig JSON: {exc}") from exc
    return config


# ---------------------------------------------------------------------------
# FantasyPros expert consensus rankings (fp) — embedded ecrData JSON
# ---------------------------------------------------------------------------

# Column layout the pipeline's 'fp' source expects (must equal COLUMN_MAPPINGS['fp']).
# SOS / ECR VS ADP are discarded by the processor after the positional rename, so they
# are emitted blank. POS holds the bare position (the pipeline strips any rank digits).
FP_OUTPUT_COLUMNS = ["ECR", "TIER", "PLAYER NAME", "TEAM", "POS", "BYE", "SOS", "ECR VS ADP"]

# Scoring format -> FantasyPros cheatsheet URL. The cheatsheet pages embed the full
# rankings as a `var ecrData = {...}` JSON blob, available even in the offseason (when
# the /rankings/*-overall.php table pages 302-redirect to consensus-cheatsheets).
FP_CHEATSHEET_URLS = {
    "ppr": "https://www.fantasypros.com/nfl/rankings/ppr-cheatsheets.php",
    "half-ppr": "https://www.fantasypros.com/nfl/rankings/half-point-ppr-cheatsheets.php",
    "standard": "https://www.fantasypros.com/nfl/rankings/consensus-cheatsheets.php",
}


def _parse_fantasypros_rankings(html: str) -> list[dict]:
    """Parse FantasyPros rankings from the embedded `ecrData` JSON into fp-schema rows.

    The cheatsheet page embeds `var ecrData = {... "players": [...] ...};`. Each player
    maps to the 8-col COLUMN_MAPPINGS['fp'] layout (SOS / ECR VS ADP left blank — the
    pipeline discards them).
    """
    match = re.search(r"var\s+ecrData\s*=\s*(\{.*?\});", html, re.DOTALL)
    if not match:
        raise RuntimeError("Could not find ecrData JSON on the FantasyPros rankings page — layout changed")
    players = json.loads(match.group(1)).get("players", [])

    rows = []
    for p in players:
        rows.append(
            {
                "ECR": p.get("rank_ecr", ""),
                "TIER": p.get("tier", ""),
                "PLAYER NAME": p.get("player_name", ""),
                "TEAM": p.get("player_team_id", ""),
                "POS": p.get("player_position_id", ""),
                "BYE": p.get("player_bye_week", ""),
                "SOS": "",
                "ECR VS ADP": "",
            }
        )
    return rows


def fetch_fantasypros_rankings(
    output_dir: str, year: int = CURRENT_SEASON, scoring: str = "ppr", min_players: int = 200
) -> str:
    """Fetch FantasyPros expert consensus rankings (fp) and save a pipeline-ready CSV.

    Reads the embedded `ecrData` JSON from the cheatsheet page (works year-round, unlike
    the /rankings/*-overall.php table which 302-redirects in the offseason) and writes the
    8-col COLUMN_MAPPINGS['fp'] layout to `FantasyPros_<year>_Draft_ALL_Rankings.csv`.

    Args:
        output_dir: Directory to save the CSV (the pipeline's update/ folder).
        year: Season year for the filename (must match FILE_MAPPINGS' fp prefix).
        scoring: One of 'ppr' (default), 'half-ppr', 'standard'.
        min_players: Coverage floor — raise if fewer rows parse (layout drift guard).

    Returns:
        Path to the saved CSV file.

    Raises:
        ValueError: if `scoring` is unknown.
        RuntimeError: if the ecrData blob is missing or fewer than `min_players` parse.
    """
    if scoring not in FP_CHEATSHEET_URLS:
        raise ValueError(f"Unknown scoring {scoring!r}; choose from {sorted(FP_CHEATSHEET_URLS)}")

    response = requests.get(FP_CHEATSHEET_URLS[scoring], headers={"User-Agent": USER_AGENT}, timeout=30)
    response.raise_for_status()

    players = _parse_fantasypros_rankings(response.text)
    if len(players) < min_players:
        raise RuntimeError(
            f"Only {len(players)} players parsed (expected >= {min_players}); "
            "FantasyPros rankings page layout may have changed"
        )

    filename = f"FantasyPros_{year}_Draft_ALL_Rankings.csv"
    output_path = os.path.join(output_dir, filename)

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FP_OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(players)

    print(f"FP rankings fetched: {len(players)} players saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# FantasyPros weekly leaders (ff-stats weekly input) — embedded reportConfig JSON
# ---------------------------------------------------------------------------

# Column layout of `data/fpts historical/weekly_data.csv`, the --weekly-data input to ff-stats.
# Week columns are the bare numbers '1'..'18'; a bye is the literal string 'BYE' (both here and
# in the source, so it passes through as-is). SEASON is appended by this fetcher — it is NOT in
# the legacy hand-downloaded file, which is exactly why a season mismatch used to go unnoticed.
WEEKLY_LEADERS_WEEKS = [str(w) for w in range(1, 19)]
WEEKLY_LEADERS_COLUMNS = ["#", "PLAYER NAME", "POS", "TEAM", *WEEKLY_LEADERS_WEEKS, "AVG", "TOTAL", "SEASON"]

# Scoring -> FantasyPros weekly-leaders report. Half-PPR is what the pipeline uses: the legacy
# weekly_data.csv matches PFR's FDPT (FanDuel = half-PPR) column exactly.
#
# Registration-fenced, so this needs the free `fp` session — the only thing that still does
# now that ADP comes from DraftSharks. Note the fence is easy to misread: `registrationFence`
# is only present (True) for ANONYMOUS visitors and is ABSENT once logged in — so "fence is
# None" means "we're authenticated", not "no fence". Judge by row count instead: anonymous
# returns an 8-row teaser vs ~734 for a full season.
WEEKLY_LEADERS_URLS = {
    "half-ppr": "https://www.fantasypros.com/nfl/reports/leaders/half-ppr.php",
    "ppr": "https://www.fantasypros.com/nfl/reports/leaders/ppr.php",
    "standard": "https://www.fantasypros.com/nfl/reports/leaders/.php",
}

# Anonymous visitors get an 8-row teaser; a logged-in session gets ~734 for a full season.
# Kept generous — this guards against the fence, not an exact row count. Used by
# _validate_fp_session as the "we're authenticated" signal.
FP_WEEKLY_LEADERS_TEASER_ROWS = 25


def _parse_fp_weekly_leaders(html: str, year: int) -> list[dict]:
    """Parse a FantasyPros weekly-leaders page into WEEKLY_LEADERS_COLUMNS rows.

    Reads the embedded `window.FP.reportConfig` blob. Fields: rank, player{name, team}, pos,
    games, wk_1..wk_18, avg, points.
    """
    config = _extract_fp_report_config(html)
    rows = config.get("table", {}).get("rows", [])

    players = []
    for row in rows:
        player = row.get("player") or {}
        name = (player.get("name") or "").strip()
        if not name:
            continue
        record = {
            "#": row.get("rank", ""),
            "PLAYER NAME": name,
            "POS": str(row.get("pos", "")).strip(),
            "TEAM": (player.get("team") or "").strip(),
        }
        for week in WEEKLY_LEADERS_WEEKS:
            # 'BYE' arrives as a literal string; anything missing stays blank.
            value = row.get(f"wk_{week}", "")
            record[week] = "" if value is None else value
        record["AVG"] = row.get("avg", "")
        record["TOTAL"] = row.get("points", "")
        record["SEASON"] = year
        players.append(record)

    return players


def fetch_fp_weekly_leaders(
    output_path: str,
    year: int,
    scoring: str = "half-ppr",
    min_players: int = 300,
) -> str:
    """Fetch a season's weekly fantasy points and write ff-stats' --weekly-data CSV.

    This replaces the hand-downloaded `weekly_data.csv`, which held a single unlabelled season.
    Writes a SEASON column so `aggregate_player_historical_stats` can verify the weekly data
    matches the season being aggregated instead of silently pairing 2025 season totals with
    2024 weekly trends.

    Requires the free FantasyPros session (`ff-rankings login fp`) — this report is
    registration-fenced like ADP, serving anonymous visitors an 8-row teaser.

    Args:
        output_path: Full path of the CSV to write (e.g. 'data/fpts historical/weekly_data.csv').
        year: Season to fetch. The page serves completed seasons via ?year=.
        scoring: One of 'half-ppr' (default — matches the pipeline), 'ppr', 'standard'.
        min_players: Coverage floor — raise if fewer rows parse. Also the backstop that catches
            the fence: a teaser lands far below any sane floor.

    Returns:
        Path to the saved CSV file.

    Raises:
        ValueError: if `scoring` is unknown.
        RuntimeError: if there's no saved session, the blob is missing, or fewer than
            `min_players` rows parse.
    """
    if scoring not in WEEKLY_LEADERS_URLS:
        raise ValueError(f"Unknown scoring {scoring!r}; choose from {sorted(WEEKLY_LEADERS_URLS)}")

    from fantasy_pipeline.scraper.auth import load_cookies

    cookies = load_cookies("fp", domain_contains="fantasypros.com")
    url = f"{WEEKLY_LEADERS_URLS[scoring]}?year={year}"
    response = requests.get(url, headers={"User-Agent": USER_AGENT}, cookies=cookies, timeout=30)
    response.raise_for_status()

    players = _parse_fp_weekly_leaders(response.text, year)
    if len(players) < min_players:
        raise RuntimeError(
            f"Only {len(players)} weekly-leader rows parsed for {year} (expected >= {min_players}); "
            "the FantasyPros session may have expired (an anonymous request gets an 8-row teaser), "
            "the season may be incomplete, or the page layout may have changed"
        )

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=WEEKLY_LEADERS_COLUMNS)
        writer.writeheader()
        writer.writerows(players)

    print(f"Weekly leaders fetched: {len(players)} players ({year}, {scoring}) saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# DraftSharks (headless-browser fetcher)
# ---------------------------------------------------------------------------

DRAFTSHARKS_URL = "https://www.draftsharks.com/rankings/half-ppr"
DS_OUTPUT_FILENAME = "rankings-half-ppr.csv"

# The exact header DraftSharks' client-side "Export Rankings" CSV emits, in order.
# This is what the pipeline renames POSITIONALLY into COLUMN_MAPPINGS['ds'] (14 cols).
DS_EXPORT_HEADER = [
    "Rank",
    "Team",
    "Player",
    "Fantasy Position",
    "Games",
    "ADP",
    "Bye",
    "SOS",
    "InjuryRisk",
    "Floor Proj",
    "Consensus Proj",
    "DS Proj",
    "CeilingProj",
    "3D Value",
]

# Pipeline-facing column order (must equal COLUMN_MAPPINGS['ds']). Used only when we
# fall back to reading the rendered DOM and must assemble the export layout ourselves.
DS_OUTPUT_COLUMNS = [
    "RK",
    "TEAM",
    "PLAYER NAME",
    "POS",
    "G",
    "DS ADP",
    "BYE",
    "SOS",
    "INJURY RISK",
    "FLOOR PROJ",
    "CONS PROJ",
    "DS PROJ",
    "CEILING PROJ",
    "3D VALUE",
]

# The export control, once logged in. The page (Alpine.js) renders two `div.export-button`
# variants and toggles them on `exportContainerOptionPrint`: one wraps a Print/Export
# dropdown, the other calls `handleExport` directly. We want the latter — the client-side
# Blob download that emits the full board. Selecting on the `@click` handler rather than a
# class keeps us on the export path and off the Print sibling.
#
# History: this export used to be reachable anonymously via `a.mobile-export-button` on a
# mobile viewport. DraftSharks removed that button; logged out, the only control left is
# `a.export-button.gated` -> /login. Hence the session requirement, and no more mobile UA.
_DS_EXPORT_SELECTOR = 'div.export-button[\\@click="handleExport"]'


def _require_playwright():
    """Import Playwright lazily; raise a friendly install hint if unavailable."""
    try:
        from playwright.sync_api import sync_playwright  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "DraftSharks fetch needs Playwright (optional 'headless' extra). Install it with:\n"
            '  uv pip install -e ".[headless]"\n'
            "  playwright install chromium"
        ) from exc
    from playwright.sync_api import sync_playwright

    return sync_playwright


def _ds_dom_row_to_output(cells: list[str]) -> dict | None:
    """Map a fully-rendered DOM ranking row (14 ordered cells) to the pipeline schema.

    Pure helper so the DOM-fallback path is unit-testable without a browser. The DOM
    table mirrors the export's column order, so cells map positionally onto
    DS_OUTPUT_COLUMNS. Returns None for non-player rows (e.g. a non-numeric rank).
    """
    cells = [c.strip() for c in cells]
    if len(cells) < len(DS_OUTPUT_COLUMNS):
        return None
    if not cells[0].lstrip().rstrip(".").isdigit():
        return None
    return dict(zip(DS_OUTPUT_COLUMNS, cells[: len(DS_OUTPUT_COLUMNS)]))


def _ds_capture_export_csv(output_path: str, storage_state: str) -> int:
    """Drive a logged-in headless browser to click Export and save its CSV.

    Returns the number of data rows written. Requires a saved session: DraftSharks
    gates the export behind login (see _DS_EXPORT_SELECTOR).
    """
    sync_playwright = _require_playwright()
    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(headless=True)
        except Exception as exc:  # browser binary missing
            raise RuntimeError(
                "Could not launch Chromium. Install the browser with:\n  playwright install chromium"
            ) from exc
        try:
            context = browser.new_context(accept_downloads=True, storage_state=storage_state)
            page = context.new_page()
            page.goto(DRAFTSHARKS_URL, wait_until="domcontentloaded", timeout=60000)

            export_btn = page.locator(_DS_EXPORT_SELECTOR)
            try:
                # 'visible', not merely 'attached': logged out, the handleExport variant is
                # absent and its hidden Print-dropdown sibling would satisfy 'attached'.
                export_btn.wait_for(state="visible", timeout=30000)
            except Exception as exc:
                raise RuntimeError(
                    "DraftSharks' Export control isn't available — the session has probably expired "
                    "(logged out, the export is gated behind /login).\nRe-authenticate with:\n"
                    "  ff-rankings login ds"
                ) from exc

            with page.expect_download(timeout=30000) as download_info:
                export_btn.click()
            download = download_info.value
            download.save_as(output_path)

            from fantasy_pipeline.scraper.auth import save_context_state

            save_context_state(context, "ds")  # sliding session — capture rotated cookies
        finally:
            browser.close()

    with open(output_path, newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise RuntimeError("DraftSharks export produced an empty CSV")
    header = [h.strip() for h in rows[0]]
    if header != DS_EXPORT_HEADER:
        raise RuntimeError(
            "DraftSharks export header changed — expected "
            f"{DS_EXPORT_HEADER}, got {header}. The export layout may have changed."
        )
    return len(rows) - 1


def fetch_draftsharks(output_dir: str, min_players: int = 150) -> str:
    """Fetch DraftSharks half-PPR rankings via a logged-in headless browser and save a CSV.

    DraftSharks' rankings page is a JS-rendered SPA: the DOM renders only ~25 players with
    no projections, so the full board is only obtainable through the page's own client-side
    "Export" button (`handleExport`). This drives that button and captures the resulting Blob
    download — the exact 14-column layout the pipeline consumes (renamed positionally into
    COLUMN_MAPPINGS['ds']).

    Requires a saved session (`ff-rankings login ds`). This fetcher previously needed no
    account, reaching an ungated mobile-only export button; DraftSharks removed it.

    Args:
        output_dir: Directory to save the CSV (the pipeline's update/ folder).
        min_players: Coverage floor — raise if fewer rows are captured (the full
            board is ~550+, so this guards against silent breakage).

    Returns:
        Path to the saved CSV file ('rankings-half-ppr.csv').

    Raises:
        RuntimeError: if there's no saved session, Playwright/Chromium is unavailable,
            the export header drifts, or fewer than `min_players` rows are captured.
    """
    from fantasy_pipeline.scraper.auth import load_storage_state

    output_path = os.path.join(output_dir, DS_OUTPUT_FILENAME)
    row_count = _ds_capture_export_csv(output_path, load_storage_state("ds"))

    if row_count < min_players:
        raise RuntimeError(
            f"Only {row_count} DraftSharks players captured (expected >= {min_players}); "
            "the export may be gated or the page layout may have changed"
        )

    print(f"DraftSharks fetched: {row_count} players saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# DraftSharks ADP (adp) — platform-specific ADP, saved-session
# ---------------------------------------------------------------------------

# Column layout the pipeline's 'adp' source expects (must equal COLUMN_MAPPINGS['adp']).
# The rename is POSITIONAL, so order is what matters. BYE / RT are placeholders the
# processor discards (they're not in STANDARD_OUTPUT_COLUMNS) — DraftSharks' export has
# no bye column, so BYE is emitted blank rather than faked.
ADP_OUTPUT_COLUMNS = ["PLAYER NAME", "TEAM", "BYE", "POS", "ADP", "MARKET INDEX", "RT"]

# The ADP board this pipeline drafts against: Sleeper, redraft, half-PPR, 12-team.
# DS_ADP_TEAMS is load-bearing twice over — it selects the board AND converts DraftSharks'
# round.pick notation to an overall pick (see _round_pick_to_overall). It also has to stay
# 12 to agree with BaseProcessor's `ADP ROUND = (ADP - 1) // 12 + 1`, which is hardcoded.
DS_ADP_SOURCE = "sleeper"
DS_ADP_SCORING = "half-ppr"
DS_ADP_TEAMS = 12

DS_ADP_EXPORT_URL = "https://www.draftsharks.com/adp/export"

# The export control on the ADP page. Logged out this renders as `a.adp__export-btn.gated`
# pointing at /login; logged in, its href carries the selected board's ids.
_DS_ADP_EXPORT_SELECTOR = "a.adp__export-btn"

# Position of each field in DraftSharks' ADP export, which is:
#   "Player Name", "Player Team", "Player Position", <adp1_name>, "Market Index 1",
# Read POSITIONALLY: the 4th header is not a fixed label but an echo of the `adp1_name`
# query param we send, so it carries no information about what the server actually returned.
_DS_ADP_NAME_IDX, _DS_ADP_TEAM_IDX, _DS_ADP_POS_IDX, _DS_ADP_ADP_IDX, _DS_ADP_MI_IDX = 0, 1, 2, 3, 4
_DS_ADP_FIXED_HEADERS = ["Player Name", "Player Team", "Player Position"]

# Positions the pipeline actually drafts. DraftSharks' board also carries aggregate rows
# that are NOT players — notably 31 `TQB` (team-QB) rows, each named after a team and so
# sharing its defense's name ('Detroit Lions' appears as both TQB @140 and DEF @267).
# They match nothing today (player_key_dict holds no team entries), but a name claimed by
# two ADP rows is a live duplicate-row bug the moment one does: the board left-merges each
# source on PLAYER ID, so a doubled ID doubles the row — the JaTavion Sanders failure.
# Drop them at the source rather than leave the landmine armed.
DS_ADP_FANTASY_POSITIONS = frozenset({"QB", "RB", "WR", "TE", "K", "DEF", "DST"})


def ds_adp_page_url(source: str = DS_ADP_SOURCE, scoring: str = DS_ADP_SCORING, teams: int = DS_ADP_TEAMS) -> str:
    """The DraftSharks ADP page for a platform/scoring/league-size combination."""
    return f"https://www.draftsharks.com/adp/{scoring}/{source}/{teams}"


def _round_pick_to_overall(value: str, teams: int) -> Optional[int]:
    """Convert DraftSharks' round.pick ADP ('2.3') to an overall pick number (15).

    DraftSharks publishes ADP as a draft slot, NOT as a number that means anything
    arithmetically: '1.10' is round 1 pick 10 (overall 10), yet reads as 1.1 to float() —
    colliding with '1.1' and sorting below '1.2'. The pipeline needs an overall pick,
    because it does `ADP Delta = ADP - avg_RK` and `ADP ROUND = (ADP - 1) // 12 + 1`.

    Returns None for a blank/unparseable cell (the caller drops the row) rather than
    guessing, so a format change surfaces as missing coverage instead of wrong numbers.

    Verified against the ADP page's own `overall_pick_number` field for all 318 players of
    the Sleeper 12-team half-PPR board: 0 mismatches.
    """
    match = re.match(r"^\s*(\d+)\.(\d+)\s*$", value or "")
    if not match:
        return None
    rnd, pick = int(match.group(1)), int(match.group(2))
    if not 1 <= pick <= teams:
        # Pick outside 1..teams means this isn't round.pick in a `teams`-sized league.
        raise RuntimeError(
            f"DraftSharks ADP '{value}' has a pick number outside 1..{teams} — the board's "
            f"league size disagrees with teams={teams}, so round.pick cannot be converted."
        )
    return (rnd - 1) * teams + pick


def _parse_ds_adp_export(text: str, teams: int) -> list[dict]:
    """Parse DraftSharks' ADP export CSV into ADP_OUTPUT_COLUMNS rows.

    Converts round.pick to an overall pick; rows without a usable ADP are dropped.
    """
    rows = list(csv.reader(io.StringIO(text), skipinitialspace=True))
    if not rows:
        raise RuntimeError("DraftSharks ADP export was empty")

    header = [h.strip() for h in rows[0]]
    if header[: len(_DS_ADP_FIXED_HEADERS)] != _DS_ADP_FIXED_HEADERS:
        raise RuntimeError(
            f"DraftSharks ADP export layout changed — expected it to start with {_DS_ADP_FIXED_HEADERS}, got {header}"
        )

    players = []
    for row in rows[1:]:
        if len(row) <= _DS_ADP_MI_IDX:
            continue
        name = row[_DS_ADP_NAME_IDX].strip()
        pos = row[_DS_ADP_POS_IDX].strip()
        adp = _round_pick_to_overall(row[_DS_ADP_ADP_IDX], teams)
        if not name or adp is None:
            continue
        if pos.upper() not in DS_ADP_FANTASY_POSITIONS:
            continue  # team aggregates (TQB), not draftable players — see the constant
        players.append(
            {
                "PLAYER NAME": name,
                "TEAM": row[_DS_ADP_TEAM_IDX].strip(),
                "BYE": "",  # not in the export; discarded downstream anyway
                "POS": pos,
                "ADP": adp,
                "MARKET INDEX": row[_DS_ADP_MI_IDX].strip(),
                "RT": "",
            }
        )

    return players


def _assert_ds_adp_board(adp_name: str, source: str, scoring: str) -> None:
    """Raise unless DraftSharks' own label for the exported column names the board we asked for.

    This is the only trustworthy provenance we get, and it must be read off the PAGE (where
    the app resolves /adp/<scoring>/<source>/<teams> into ids), never off the export: the
    export endpoint echoes back whatever `adp1_name` we hand it, answers a wrong id with a
    full, plausible board for a DIFFERENT platform, and answers an unknown id with a bare
    header and HTTP 200. Nothing about its response says which platform it actually is.
    """
    label = adp_name.lower()
    if source.lower() not in label:
        raise RuntimeError(
            f"DraftSharks' export is labelled {adp_name!r}, which does not name '{source}'. "
            "Refusing to save it — the platform ids on the ADP page may have been renumbered."
        )
    # 'half-ppr' is written '0.5 PPR' in DraftSharks' label.
    expected_scoring = "0.5 ppr" if scoring == "half-ppr" else scoring.replace("-", " ")
    if expected_scoring not in label:
        raise RuntimeError(
            f"DraftSharks' export is labelled {adp_name!r}, which does not name '{expected_scoring}' "
            "scoring. Refusing to save it — the page's format ids may have changed."
        )


def _read_ds_adp_export_params(page_url: str, storage_state: str) -> tuple[dict, str]:
    """Return the ADP export's query params and DraftSharks' label for the selected board.

    The page URL (/adp/<scoring>/<source>/<teams>) is the source of truth for WHICH board is
    meant, but the numeric `format::source::size` ids it resolves to are computed client-side
    and appear only in the rendered export link — so this reads them from the live page rather
    than hardcoding ids that would silently point at another platform if DraftSharks renumbers.
    """
    from urllib.parse import parse_qs, urlparse

    sync_playwright = _require_playwright()
    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(headless=True)
        except Exception as exc:  # browser binary missing
            raise RuntimeError(
                "Could not launch Chromium. Install the browser with:\n  playwright install chromium"
            ) from exc
        try:
            context = browser.new_context(storage_state=storage_state)
            page = context.new_page()
            page.goto(page_url, wait_until="domcontentloaded", timeout=60000)

            link = page.locator(_DS_ADP_EXPORT_SELECTOR).first
            try:
                link.wait_for(state="visible", timeout=30000)
            except Exception as exc:
                raise RuntimeError(
                    f"DraftSharks' ADP export link never appeared on {page_url} — the page layout may have changed."
                ) from exc

            href = link.get_attribute("href") or ""
            from fantasy_pipeline.scraper.auth import save_context_state

            save_context_state(context, "ds")  # sliding session — capture rotated cookies
        finally:
            browser.close()

    # Logged out, the control is `a.adp__export-btn.gated` -> /login.
    if "/adp/export" not in href:
        raise RuntimeError(
            "DraftSharks' ADP export is gated — the 'ds' session has expired or is missing.\n"
            "Re-authenticate with:\n  ff-rankings login ds"
        )

    query = parse_qs(urlparse(href).query)
    adp1 = (query.get("adp1") or [""])[0]
    adp1_name = (query.get("adp1_name") or [""])[0]
    if not adp1:
        raise RuntimeError(f"DraftSharks' ADP export link carried no 'adp1' board id: {href!r}")

    return {"adp1": adp1, "adp1_name": adp1_name, "sort_column": "adp1"}, adp1_name


def fetch_draftsharks_adp(
    output_dir: str,
    year: int = CURRENT_SEASON,
    source: str = DS_ADP_SOURCE,
    scoring: str = DS_ADP_SCORING,
    teams: int = DS_ADP_TEAMS,
    min_players: int = 200,
) -> str:
    """Fetch platform-specific ADP from DraftSharks and save a pipeline-ready CSV.

    Defaults to the Sleeper 12-team half-PPR redraft board
    (https://www.draftsharks.com/adp/half-ppr/sleeper/12), replacing the FantasyPros
    consensus ADP this pipeline used to draft against. Writes the 7-col
    COLUMN_MAPPINGS['adp'] layout to `DraftSharks_<year>_<Source>_ADP.csv`.

    Two things make this less direct than it looks:

    1. **ADP arrives as round.pick, not an overall pick.** '1.10' is overall 10 but floats to
       1.1. Everything downstream (`ADP Delta`, `ADP ROUND`, `POS ADP`) treats ADP as an
       overall pick, so the values are converted (see _round_pick_to_overall).
    2. **The export endpoint has no provenance.** It labels its ADP column with the
       `adp1_name` we send it, serves a full board for the wrong platform if the ids are
       wrong, and returns HTTP 200 + a bare header for ids it doesn't know. So the board ids
       are read off the live page (where the app resolves the URL's platform into ids) and
       DraftSharks' own label is asserted to name the requested platform before anything is
       written.

    Requires a saved session (`ff-rankings login ds`) — the same one `fetch_draftsharks`
    uses; the ADP export is gated behind /login.

    Args:
        output_dir: Directory to save the CSV (the pipeline's update/ folder).
        year: Season year for the filename (the page always serves the live season).
        source: ADP platform slug as it appears in the page URL (e.g. 'sleeper', 'espn').
        scoring: Scoring slug as it appears in the page URL (e.g. 'half-ppr', 'ppr').
        teams: League size. Selects the board AND converts round.pick to an overall pick.
        min_players: Coverage floor — raise if fewer rows parse (layout drift guard).

    Returns:
        Path to the saved CSV file.

    Raises:
        RuntimeError: if there's no saved session, Playwright/Chromium is unavailable, the
            export is gated, the board's label doesn't name `source`/`scoring`, the export
            layout drifts, or fewer than `min_players` rows parse.
    """
    from fantasy_pipeline.scraper.auth import load_cookies, load_storage_state

    page_url = ds_adp_page_url(source, scoring, teams)
    params, adp_name = _read_ds_adp_export_params(page_url, load_storage_state("ds"))
    _assert_ds_adp_board(adp_name, source, scoring)

    cookies = load_cookies("ds", domain_contains="draftsharks")
    response = requests.get(
        DS_ADP_EXPORT_URL, params=params, headers={"User-Agent": USER_AGENT}, cookies=cookies, timeout=60
    )
    response.raise_for_status()

    players = _parse_ds_adp_export(response.text, teams)
    if len(players) < min_players:
        raise RuntimeError(
            f"Only {len(players)} ADP players parsed from {adp_name!r} (expected >= {min_players}); "
            "the 'ds' session may have expired, or the export may have changed"
        )

    filename = f"DraftSharks_{year}_{source.capitalize()}_ADP.csv"
    output_path = os.path.join(output_dir, filename)

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=ADP_OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(players)

    print(f"ADP fetched: {len(players)} players from {adp_name!r} saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# PFF (paywalled — saved-session headless fetcher)
# ---------------------------------------------------------------------------

PFF_RANKINGS_URL = "https://www.pff.com/fantasy/rankings/draft"

# The 9-col header row inside PFF's "Export"/"Download" CSV (the file the manual
# workflow already drops as Draft-rankings-export-<year>.csv). The pipeline finds this
# header row (the file has a title row above it) and renames POSITIONALLY into
# COLUMN_MAPPINGS['pff'].
PFF_EXPORT_HEADER = [
    "Overall Rank",
    "Full Name",
    "Team Abbreviation",
    "Position",
    "Position Rank",
    "Bye Week",
    "ADP",
    "Projected Points",
    "Auction Value",
]


# PFF's CSV export control. `data-testid` is the stable hook — the accessible name is
# "Download CSV" while the visible text is just "CSV", so name/text matching is brittle
# (and matched the *locked* button, producing a bare download timeout).
PFF_CSV_BUTTON_SELECTOR = 'button[data-testid="csvDownloadButton"]'


def _pff_output_filename(year: int) -> str:
    """Filename matching FILE_MAPPINGS' 'Draft-rankings-export' prefix."""
    return f"Draft-rankings-export-{year}.csv"


def _pff_export_is_locked(button) -> bool:
    """True if PFF's CSV button renders a lock icon (account lacks the export entitlement).

    An un-entitled (but logged-in) account still gets the button; it just wears a lock and
    redirects to /subscribe on click. Entitled, the same slot holds a download icon.
    """
    try:
        return bool(button.locator('[data-testid*="lockIcon"], [class*="lockIcon"]').count())
    except Exception:
        return False


def _click_pff_export(page) -> None:
    """Click PFF's CSV export control, failing loudly if the account isn't entitled."""
    button = page.locator(PFF_CSV_BUTTON_SELECTOR).first
    try:
        button.wait_for(state="visible", timeout=15000)
    except Exception as exc:
        raise RuntimeError(
            f"Could not find PFF's CSV export control ({PFF_CSV_BUTTON_SELECTOR}). "
            "The session may have expired, or the page layout changed."
        ) from exc

    if _pff_export_is_locked(button):
        raise RuntimeError(
            "PFF's CSV export is locked for this account — it redirects to /subscribe.\n"
            "The saved session is logged in but lacks the CSV entitlement. If your "
            "subscription is active, refresh the session with:\n  ff-rankings login pff"
        )

    button.click()


def _pff_capture_export_csv(output_path: str, storage_state: str) -> int:
    """Drive a logged-in headless browser to export PFF rankings; save the CSV.

    Reuses the saved session (`storage_state`) so no password is handled here.
    Returns the number of data rows written (rows after the 'Overall Rank' header).
    """
    sync_playwright = _require_playwright()
    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(headless=True)
        except Exception as exc:  # browser binary missing
            raise RuntimeError(
                "Could not launch Chromium. Install the browser with:\n  playwright install chromium"
            ) from exc
        try:
            context = browser.new_context(
                accept_downloads=True,
                storage_state=storage_state,
            )
            page = context.new_page()
            page.goto(PFF_RANKINGS_URL, wait_until="domcontentloaded", timeout=60000)

            with page.expect_download(timeout=30000) as download_info:
                _click_pff_export(page)
            download = download_info.value
            download.save_as(output_path)
            from fantasy_pipeline.scraper.auth import save_context_state

            save_context_state(context, "pff")  # sliding session — capture rotated cookies
        finally:
            browser.close()

    return _validate_pff_csv(output_path)


def _validate_pff_csv(output_path: str) -> int:
    """Validate the saved PFF CSV's header and return its data-row count.

    The export carries a title row above the real header, so we locate the
    'Overall Rank' row and validate the 9 columns there.
    """
    with open(output_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise RuntimeError("PFF export produced an empty CSV")

    header_idx = next(
        (i for i, r in enumerate(rows) if r and r[0].strip() == "Overall Rank"),
        None,
    )
    if header_idx is None:
        raise RuntimeError(
            "PFF export missing the 'Overall Rank' header row — not logged in, or the export layout changed"
        )
    header = [h.strip() for h in rows[header_idx]]
    if header != PFF_EXPORT_HEADER:
        raise RuntimeError(f"PFF export header changed — expected {PFF_EXPORT_HEADER}, got {header}")
    data_rows = [r for r in rows[header_idx + 1 :] if r and r[0].strip()]
    return len(data_rows)


def fetch_pff(output_dir: str, year: int = CURRENT_SEASON, min_players: int = 200) -> str:
    """Fetch PFF draft rankings via a saved logged-in session and save a CSV.

    PFF's rankings are behind a premium subscription. This reuses the session saved by
    `ff-rankings login pff` (no password handled here) to drive the page's own
    Export/Download, capturing the exact CSV the pipeline already consumes
    (renamed positionally into COLUMN_MAPPINGS['pff']) → Draft-rankings-export-<year>.csv.

    Args:
        output_dir: Directory to save the CSV (the pipeline's update/ folder).
        year: Season year for the filename (matches FILE_MAPPINGS' pff prefix).
        min_players: Coverage floor — raise if fewer rows are captured.

    Returns:
        Path to the saved CSV file.

    Raises:
        RuntimeError: if there is no saved session, Playwright/Chromium is unavailable,
            the export header drifts, or fewer than `min_players` rows are captured.
    """
    from fantasy_pipeline.scraper.auth import load_storage_state

    storage_state = load_storage_state("pff")
    output_path = os.path.join(output_dir, _pff_output_filename(year))
    row_count = _pff_capture_export_csv(output_path, storage_state)

    if row_count < min_players:
        raise RuntimeError(
            f"Only {row_count} PFF players captured (expected >= {min_players}); "
            "the session may have expired or the export may be gated"
        )

    print(f"PFF fetched: {row_count} players saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# FantasyPoints / Scott Barrett (paywalled — saved-session headless fetcher)
# ---------------------------------------------------------------------------

# Redraft rankings SPA. The page defaults to Hansen's rankings; a "BARRETT'S RANKINGS"
# tab switches to Scott Barrett's board (the one the manual workflow uses). Overridable
# via --url for live verification.
FPTS_RANKINGS_URL = "https://www.fantasypoints.com/nfl/rankings/redraft"

# The 7-col header FantasyPoints' "Download as CSV" emits (the file the manual workflow
# drops as "Scott Barrett*.csv"). The pipeline renames POSITIONALLY into COLUMN_MAPPINGS['fpts'].
FPTS_EXPORT_HEADER = ["Overall", "NAME", "POS", "TEAM", "BYE", "TIER", "EXODIA"]


def _fpts_output_filename(year: int) -> str:
    """Filename matching FILE_MAPPINGS' 'Scott Barrett' prefix."""
    return f"Scott Barrett {year} Redraft Rankings.csv"


def _select_fpts_barrett(page) -> None:
    """Switch the redraft rankings SPA from Hansen's board to Scott Barrett's.

    The page loads Hansen's rankings by default; clicking the "BARRETT'S RANKINGS" tab
    re-renders the table with Barrett's data (page title gains "Barrett's").
    """
    tab = page.locator('a:has-text("BARRETT\'S RANKINGS")').first
    tab.wait_for(state="visible", timeout=20000)
    tab.click()
    # Confirm the board actually switched before exporting (guards against silently
    # downloading Hansen's rankings under the Barrett filename).
    try:
        page.wait_for_function("() => /barrett/i.test(document.title)", timeout=15000)
    except Exception as exc:
        raise RuntimeError(
            "Clicked 'BARRETT'S RANKINGS' but the page title never switched to Barrett's "
            "board — the rankings SPA may have changed."
        ) from exc


# The on-page rankings heading, e.g. "Scott Barrett's 2025 NFL Redraft Rankings".
# This is the ONLY trustworthy season signal on the page — see _assert_fpts_season.
_FPTS_HEADING_SELECTOR = "h1"
_FPTS_HEADING_SEASON_RE = re.compile(r"(20\d\d)\s+NFL\s+Redraft\s+Rankings", re.I)


def _assert_fpts_season(page, year: int) -> None:
    """Raise unless the rendered board is actually `year`'s rankings.

    **Do not trust `document.title` for the season.** FantasyPoints templates it to the
    *current* year while the body still serves the previous season's board: in the 2026
    preseason the title read "Scott Barrett's 2026 Redraft Fantasy Football Rankings"
    above an `<h1>` of "Scott Barrett's 2025 NFL Redraft Rankings" (updated 2025-08-30).
    Without this guard the fetcher silently saves last season's ranks as
    `Scott Barrett <year> Redraft Rankings.csv` — a filename that lies about its contents,
    which then quietly contaminates avg_RK. (Barrett publishes late; expect this to fire
    through the early preseason.)
    """
    try:
        heading = page.locator(_FPTS_HEADING_SELECTOR).first
        heading.wait_for(state="attached", timeout=10000)
        text = (heading.text_content() or "").strip()
    except Exception as exc:
        raise RuntimeError(
            f"Could not read the FantasyPoints rankings heading ({_FPTS_HEADING_SELECTOR}) to "
            "verify the season — the page layout may have changed."
        ) from exc

    match = _FPTS_HEADING_SEASON_RE.search(text)
    if not match:
        raise RuntimeError(
            f"Could not parse a season from the FantasyPoints heading {text!r}; refusing to "
            "save rankings that can't be confirmed as the right season."
        )

    found = int(match.group(1))
    if found != year:
        raise RuntimeError(
            f"FantasyPoints is serving Barrett's {found} board, not {year} "
            f"(heading: {text!r}). His {year} rankings are probably not published yet — "
            f"the page title claims {year} but the data is {found}'s. Re-run once they're live, "
            f"or pass --year {found} to fetch that season deliberately."
        )


def _click_fpts_csv_download(page) -> None:
    """Click the DataTables 'Download as CSV' button (fires a client-side Blob download)."""
    btn = page.locator("button.buttons-csv, button:has-text('Download as CSV')").first
    btn.wait_for(state="visible", timeout=15000)
    btn.click()


def _fpts_capture_export_csv(output_path: str, storage_state: str, rankings_url: str, year: int) -> int:
    """Drive a logged-in headless browser to export Barrett's rankings; save the CSV.

    Reuses the saved session (`storage_state`); no password handled here. Selects
    Barrett's board, verifies it is `year`'s (see _assert_fpts_season), then captures its
    "Download as CSV". Returns the data-row count.
    """
    sync_playwright = _require_playwright()
    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(headless=True)
        except Exception as exc:  # browser binary missing
            raise RuntimeError(
                "Could not launch Chromium. Install the browser with:\n  playwright install chromium"
            ) from exc
        try:
            context = browser.new_context(
                accept_downloads=True,
                storage_state=storage_state,
                viewport={"width": 1400, "height": 1000},
            )
            page = context.new_page()
            page.goto(rankings_url, wait_until="networkidle", timeout=60000)

            _select_fpts_barrett(page)
            # Verify the season BEFORE downloading — a stale board must never reach disk
            # under a filename claiming the current year.
            _assert_fpts_season(page, year)
            with page.expect_download(timeout=30000) as download_info:
                _click_fpts_csv_download(page)
            download = download_info.value
            download.save_as(output_path)
            from fantasy_pipeline.scraper.auth import save_context_state

            save_context_state(context, "fpts")  # sliding session — capture rotated cookies
        finally:
            browser.close()

    return _validate_fpts_csv(output_path)


def _validate_fpts_csv(output_path: str) -> int:
    """Validate the saved FantasyPoints CSV header (row 1) and return its data-row count."""
    with open(output_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise RuntimeError("FantasyPoints export produced an empty CSV")

    header = [h.strip() for h in rows[0]]
    if header != FPTS_EXPORT_HEADER:
        raise RuntimeError(
            f"FantasyPoints export header changed — expected {FPTS_EXPORT_HEADER}, got {header}. "
            "Not logged in, wrong page, or the export layout changed."
        )
    data_rows = [r for r in rows[1:] if r and r[0].strip()]
    return len(data_rows)


def fetch_fpts(
    output_dir: str, year: int = CURRENT_SEASON, min_players: int = 90, rankings_url: str = FPTS_RANKINGS_URL
) -> str:
    """Fetch FantasyPoints (Scott Barrett) redraft rankings via a saved session; save a CSV.

    FantasyPoints is behind a subscription. This reuses the session saved by
    `ff-rankings login fpts` (no password handled here) to drive the page's own
    Export/Download, capturing the CSV the pipeline already consumes (renamed positionally
    into COLUMN_MAPPINGS['fpts']) → "Scott Barrett <year> Redraft Rankings.csv".

    Args:
        output_dir: Directory to save the CSV (the pipeline's update/ folder).
        year: Season year for the filename.
        min_players: Coverage floor — raise if fewer rows are captured (the redraft
            board is ~100, so the default 90 catches breakage with a little margin).
        rankings_url: Rankings page to export from (overridable for live verification).

    Returns:
        Path to the saved CSV file.

    Raises:
        RuntimeError: if there is no saved session, Playwright/Chromium is unavailable,
            the export header drifts, or fewer than `min_players` rows are captured.
    """
    from fantasy_pipeline.scraper.auth import load_storage_state

    storage_state = load_storage_state("fpts")
    output_path = os.path.join(output_dir, _fpts_output_filename(year))
    row_count = _fpts_capture_export_csv(output_path, storage_state, rankings_url, year)

    if row_count < min_players:
        raise RuntimeError(
            f"Only {row_count} FantasyPoints players captured (expected >= {min_players}); "
            "the session may have expired or the export may be gated"
        )

    print(f"FantasyPoints fetched: {row_count} players saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# JJ Zachariason / LateRound (Patreon — saved-session API attachment fetcher)
# ---------------------------------------------------------------------------
#
# JJ's redraft rankings are a downloadable attachment on a Patreon post (collection
# 47664). Two live realities shape this fetcher:
#   1. Patreon post *pages* are gated by a Cloudflare Turnstile challenge that flags the
#      headless browser — but the Patreon *JSON API* (same session cookies) is not. So we
#      read attachments via the API instead of clicking the HTML.
#   2. The attachment is now a 5-col CSV (`Overall,Player,Position,Pos Rank,Tier`) — the
#      `Auction` column the old .xlsx had was dropped. We pad it back to the 6-col width
#      the pipeline renames POSITIONALLY into COLUMN_MAPPINGS['jj'] (source order already
#      matches: RK, PLAYER NAME, POS, POS RANK, TIER, AUCTION).
JJ_COLLECTION_URL = "https://www.patreon.com/collection/47664"
JJ_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

# 6-col output header; the pipeline renames it positionally into COLUMN_MAPPINGS['jj'].
JJ_OUTPUT_COLUMNS = ["Overall", "Player", "Position", "Pos Rank", "Tier", "Auction"]


def _jj_output_filename(year: int) -> str:
    """Filename matching FILE_MAPPINGS' redraft 'Redraft1QB_' prefix (CSV now, not xlsx)."""
    return f"Redraft1QB_{year}.csv"


def _jj_is_redraft_title(title: str) -> bool:
    """True if a collection post title is the 1QB redraft (not superflex/ROS/weekly)."""
    t = (title or "").lower()
    return (
        "1qb" in t and ("redraft" in t or "season-long" in t) and not re.search(r"superflex|rest-of-season|weekly", t)
    )


def _jj_post_id_from_url(url: str) -> str | None:
    """Extract the numeric Patreon post id from a post URL, or None."""
    m = re.search(r"/posts/(\d+)", url or "")
    return m.group(1) if m else None


def _jj_adapt_rows(rows: list) -> list:
    """Normalize attachment rows to the 6-col COLUMN_MAPPINGS['jj'] width.

    The current source has 5 cols (no Auction); older xlsx had 6. Pad 5→6 (blank Auction),
    truncate any extras, and drop spacer rows. Header is row 0.
    """
    width = len(JJ_OUTPUT_COLUMNS)  # 6
    out = []
    for r in rows:
        cells = ["" if c is None else c for c in r]
        if len(cells) == width - 1:
            cells = cells + [""]  # pad the dropped Auction column
        if len(cells) >= width:
            out.append([str(c) for c in cells[:width]])
        # too-short / blank spacer rows are skipped
    if len(out) < 2:
        raise RuntimeError("JJ attachment had no usable rows after adapting to 6 columns")
    return out


def _jj_data_row_count(rows: list) -> int:
    """Count data rows (after the header) whose first cell is a numeric rank."""
    return sum(1 for r in rows[1:] if r and str(r[0]).strip().isdigit())


def _jj_rows_from_attachment(file_name: str, raw: bytes) -> list:
    """Parse the downloaded attachment (.csv or .xlsx) into a list of rows."""
    name = (file_name or "").lower()
    if name.endswith(".csv"):
        return list(csv.reader(io.StringIO(raw.decode("utf-8-sig"))))
    if name.endswith(".xlsx"):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        try:
            sheet = "Rankings and Tiers" if "Rankings and Tiers" in wb.sheetnames else wb.sheetnames[-1]
            return [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        finally:
            wb.close()
    raise RuntimeError(f"Unexpected JJ attachment type: {file_name!r} (want .csv/.xlsx)")


def _jj_api_json(context, url: str) -> dict:
    """GET a Patreon JSON API URL via the session's request context; raise if blocked."""
    r = context.request.get(url, headers={"Accept": "application/json"})
    body = r.text()
    if r.status != 200 or not body.strip().startswith("{"):
        raise RuntimeError(
            f"Patreon API returned {r.status} for {url} — the session may have expired "
            "or be blocked. Re-run `ff-rankings login jj`."
        )
    return json.loads(body)


def _jj_discover_post_id(page) -> str:
    """Find the latest 1QB redraft post id from the collection page (newest-first)."""
    page.goto(JJ_COLLECTION_URL, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(4000)
    for _ in range(3):  # lazy-loaded list — scroll to surface more posts
        page.mouse.wheel(0, 4000)
        page.wait_for_timeout(1200)
    posts = page.evaluate(
        r"""() => {
          const seen = new Set(), out = [];
          document.querySelectorAll('a[href*="/posts/"]').forEach(a => {
            const m = (a.getAttribute('href') || '').match(/\/posts\/(\d+)/);
            if (!m || seen.has(m[1])) return;
            seen.add(m[1]);
            out.push({id: m[1], title: (a.innerText || '').replace(/\s+/g, ' ').trim()});
          });
          return out;
        }"""
    )
    pick = next((p for p in posts if _jj_is_redraft_title(p["title"])), None)
    if not pick:
        raise RuntimeError(
            "No 1QB redraft post found in the LateRound collection "
            f"({JJ_COLLECTION_URL}) — pass --post-url with the current post explicitly."
        )
    return pick["id"]


def _jj_attachment(context, post_id: str) -> tuple:
    """Return (file_name, download_url) of the post's Redraft1QB .csv/.xlsx attachment."""
    api = (
        f"https://www.patreon.com/api/posts/{post_id}"
        "?include=attachments_media&fields[media]=file_name,download_url,mimetype"
        "&json-api-version=1.0"
    )
    data = _jj_api_json(context, api)
    media = [m["attributes"] for m in data.get("included", []) if m.get("type") == "media"]
    target = next(
        (
            m
            for m in media
            if "redraft1qb" in (m.get("file_name") or "").lower()
            and (m.get("file_name") or "").lower().endswith((".csv", ".xlsx"))
        ),
        None,
    )
    if not target:
        raise RuntimeError(
            f"Post {post_id} has no Redraft1QB .csv/.xlsx attachment (found: {[m.get('file_name') for m in media]})"
        )
    return target["file_name"], target["download_url"]


def _jj_fetch_rows(storage_state: str, post_url: str | None) -> list:
    """Discover (or use) the post, download its Redraft1QB attachment via the API, adapt.

    Returns adapted 6-col rows (header + data). Uses the API for attachments because the
    Patreon post HTML is Cloudflare-gated for the headless browser.
    """
    sync_playwright = _require_playwright()
    with sync_playwright() as p:
        try:
            browser = p.chromium.launch(headless=True)
        except Exception as exc:  # browser binary missing
            raise RuntimeError(
                "Could not launch Chromium. Install the browser with:\n  playwright install chromium"
            ) from exc
        try:
            context = browser.new_context(
                accept_downloads=True,
                storage_state=storage_state,
                user_agent=JJ_UA,
                viewport={"width": 1400, "height": 1200},
            )
            page = context.new_page()
            post_id = _jj_post_id_from_url(post_url) if post_url else _jj_discover_post_id(page)
            if not post_id:
                raise RuntimeError(f"Could not parse a post id from --post-url {post_url!r}")
            file_name, download_url = _jj_attachment(context, post_id)
            raw = context.request.get(download_url).body()
            from fantasy_pipeline.scraper.auth import save_context_state

            save_context_state(context, "jj")  # sliding session — capture rotated cookies
        finally:
            browser.close()

    return _jj_adapt_rows(_jj_rows_from_attachment(file_name, raw))


def fetch_jj(
    output_dir: str, post_url: Optional[str] = None, year: int = CURRENT_SEASON, min_players: int = 150
) -> str:
    """Fetch JJ Zachariason's 1QB redraft rankings from Patreon via a saved session.

    Reuses the session from `ff-rankings login jj` (no password handled). By default it
    auto-discovers the latest 1QB redraft post in the LateRound collection; pass
    `post_url` to target a specific post. The attachment is downloaded through the Patreon
    JSON API (the post HTML is Cloudflare-gated), adapted to the 6-col COLUMN_MAPPINGS['jj']
    width (the current source dropped the Auction column), and written as a CSV.

    Args:
        output_dir: Directory to save the CSV (the pipeline's update/ folder).
        post_url: Optional Patreon post URL; if omitted, the latest 1QB redraft post is found.
        year: Season year for the filename.
        min_players: Coverage floor — raise if fewer rows are found (the board is ~250).

    Returns:
        Path to the saved CSV file ('Redraft1QB_<year>.csv').

    Raises:
        RuntimeError: if there is no saved session, Playwright/Chromium is unavailable, the
            session is blocked, no redraft post/attachment is found, or coverage is low.
    """
    from fantasy_pipeline.scraper.auth import load_storage_state

    storage_state = load_storage_state("jj")
    rows = _jj_fetch_rows(storage_state, post_url)
    count = _jj_data_row_count(rows)
    if count < min_players:
        raise RuntimeError(
            f"Only {count} JJ players found (expected >= {min_players}); "
            "the session may have expired or the wrong attachment was downloaded"
        )

    output_path = os.path.join(output_dir, _jj_output_filename(year))
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)

    print(f"JJ fetched: {count} players saved to {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Session validation + auto-login (paywalled sources)
# ---------------------------------------------------------------------------
#
# A cheap, source-specific "are we still logged in?" probe so the CLI can pop the login
# window *only* when a session is actually expired — instead of failing mid-fetch.


def _validate_pff_session(context) -> bool:
    """True if PFF's CSV export is present *and unlocked* (i.e. logged in and entitled).

    The presence of the button is not enough: an un-entitled account still renders it,
    carrying a lock icon and redirecting to /subscribe on click. Checking only for the
    control reports a session as valid while every fetch fails on a download timeout, so
    the lock is the signal that matters.
    """
    page = context.new_page()
    try:
        page.goto(PFF_RANKINGS_URL, wait_until="domcontentloaded", timeout=40000)
        page.wait_for_timeout(3000)
        button = page.locator(PFF_CSV_BUTTON_SELECTOR)
        button.first.wait_for(state="attached", timeout=8000)
        return not _pff_export_is_locked(button.first)
    except Exception:
        return False
    finally:
        page.close()


def _validate_fpts_session(context) -> bool:
    """True if the FantasyPoints redraft page renders its 'Download as CSV' control."""
    page = context.new_page()
    try:
        page.goto(FPTS_RANKINGS_URL, wait_until="networkidle", timeout=40000)
        page.wait_for_timeout(2500)
        page.locator("button.buttons-csv, button:has-text('Download as CSV')").first.wait_for(
            state="attached", timeout=8000
        )
        return True
    except Exception:
        return False
    finally:
        page.close()


def _validate_jj_session(context) -> bool:
    """True if Patreon's current-user API reports a logged-in user."""
    try:
        data = _jj_api_json(context, "https://www.patreon.com/api/current_user?json-api-version=1.0")
        return bool(data.get("data", {}).get("id"))
    except Exception:
        return False


def _validate_fp_session(context) -> bool:
    """True if FantasyPros' weekly-leaders report returns a full season, not a fenced teaser.

    Probes the weekly-leaders report because that is the only thing the `fp` session still
    gates (`ff-stats fetch-weekly`): `fetch-fp` reads the un-fenced cheatsheet, and ADP now
    comes from DraftSharks. Validate what the session is actually used for.

    Judge by row count, not by the `registrationFence` flag: that key is only present for
    ANONYMOUS visitors and vanishes once logged in, so `fence is None` means authenticated.

    Uses the context's request API (which carries its cookies) — no page render needed,
    since the report is server-rendered into the HTML.
    """
    try:
        response = context.request.get(f"{WEEKLY_LEADERS_URLS['half-ppr']}?year={LAST_COMPLETED_SEASON}", timeout=30000)
        if not response.ok:
            return False
        config = _extract_fp_report_config(response.text())
        return len(config.get("table", {}).get("rows", [])) > FP_WEEKLY_LEADERS_TEASER_ROWS
    except Exception:
        return False


def _validate_ds_session(context) -> bool:
    """True if DraftSharks' rankings page shows the ungated `handleExport` control.

    Logged out, that variant isn't rendered at all — only `a.export-button.gated` -> /login.
    """
    page = context.new_page()
    try:
        page.goto(DRAFTSHARKS_URL, wait_until="domcontentloaded", timeout=40000)
        page.wait_for_timeout(3000)
        page.locator(_DS_EXPORT_SELECTOR).first.wait_for(state="visible", timeout=8000)
        return True
    except Exception:
        return False
    finally:
        page.close()


_SESSION_VALIDATORS = {
    "fp": _validate_fp_session,
    "ds": _validate_ds_session,
    "pff": _validate_pff_session,
    "fpts": _validate_fpts_session,
    "jj": _validate_jj_session,
}


def validate_session(source: str) -> bool:
    """Return True if ``source``'s saved session still authenticates (cheap live probe).

    Returns False if there's no saved session, the source isn't validatable, or the probe
    fails for any reason (treated as 'needs login').
    """
    from fantasy_pipeline.scraper.auth import storage_state_path

    if not storage_state_path(source).exists():
        return False
    validator = _SESSION_VALIDATORS.get(source)
    if validator is None:
        return False

    sync_playwright = _require_playwright()
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            try:
                context = browser.new_context(
                    storage_state=str(storage_state_path(source)),
                    user_agent=JJ_UA,
                )
                return validator(context)
            finally:
                browser.close()
    except Exception:
        return False


def ensure_session(source: str, auto_login: bool = True, timeout_minutes: int = 10) -> bool:
    """Ensure ``source`` has a valid session, opening the login window if needed.

    Probes the saved session; if it's missing/expired and ``auto_login`` is set, opens the
    headed login window (you finish the login), then re-validates. Returns whether the
    session is valid afterward.
    """
    if validate_session(source):
        return True
    if not auto_login:
        return False

    from fantasy_pipeline.scraper.auth import login

    print(f"\n🔑 '{source}' session is missing or expired — opening a login window...")
    try:
        login(source, timeout_minutes=timeout_minutes)
    except Exception as e:
        print(f"   ⚠️  Login did not complete: {e}")
        return False
    return validate_session(source)
